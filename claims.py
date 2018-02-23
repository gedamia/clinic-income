import xlrd
import openpyxl
import sys
import os
import re
from openpyxl import load_workbook
#import pdb
#import ipdb
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell



import logging
import logging.handlers 
'''logging.basicConfig(level=logging.DEBUG,
                format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                datefmt='%a, %d %b %Y %H:%M:%S',
                filename='myapp.log',
                filemode='w')'''
    


from openpyxl.styles import numbers, is_date_format


CLAIMS_PATH = r"C:\Evelyn\report\owe money.xlsx"
ACCOUNTING_PATH = r"C:\Evelyn\accounting"
MONEY_PATH = ACCOUNTING_PATH+r"\money patient"
REGULAR_PATH = ACCOUNTING_PATH+r"\regular patient"


LOG_FILE = 'claims_log.txt'

'''def search_file(path, last_name, first_name, birth):
    for filename in os.listdir(path):
        fp = os.path.join(path, filename)
        if not os.path.isfile(fp):
            #return search_file(fp, last_name, first_name, birth)
            search_file(fp, last_name, first_name, birth)
        else:
            if bool( re.search(last_name, filename, re.IGNORECASE) and re.search(first_name, filename, re.IGNORECASE) ):
                #pdb.set_trace()
                return fp
            else:
                continue
    return fp'''

def search_in_money_patient(last_name, first_name, birth):
    for filename in os.listdir(MONEY_PATH):
        fp = os.path.join(MONEY_PATH, filename)
        if os.path.isfile(fp):
            if bool( re.search(last_name, filename, re.IGNORECASE) and re.search(first_name, filename, re.IGNORECASE) ):
                return fp
            else:
                continue

def search_in_regular_patient(last_name, first_name, birth):
    for filename in os.listdir(REGULAR_PATH):
        fp = os.path.join(REGULAR_PATH, filename)
        if os.path.isfile(fp):
            if bool( re.search(last_name, filename, re.IGNORECASE) and re.search(first_name, filename, re.IGNORECASE) ):
                return fp
            else:
                continue

def append_claim(fp,v):
    wb = openpyxl.load_workbook(fp)
    name_list = wb.get_sheet_names()
    my_sheet = wb.get_sheet_by_name(name_list[0])
    my_sheet.append(v)
    wb.save(fp)

data = xlrd.open_workbook(CLAIMS_PATH)

claims_table = data.sheets()[0]

for i in range(claims_table.nrows):
    last_name = claims_table.cell(i,3).value
    first_name = claims_table.cell(i,4).value
    birth = claims_table.cell(i,5).value
    fp = search_in_regular_patient(last_name, first_name, birth)
    # in the regular patient
    if  fp:
        append_claim(fp, claims_table.row_values(i))
    else:
        fp = search_in_money_patient(last_name, first_name, birth)
        #in the money patient
        if fp:
            append_claim(fp, claims_table.row_values(i))
        else:
            #create new file in the accounting path
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            # append data
            new_ws.append(claims_table.row_values(i))
            new_wb.save(REGULAR_PATH+"\\"+last_name+","+first_name+".xlsx")
