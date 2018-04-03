import xlrd
import openpyxl
import sys
import os
import os.path
import re
import time
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
import string

import shutil

import logging
import logging.handlers 
'''logging.basicConfig(level=logging.DEBUG,
                format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                datefmt='%a, %d %b %Y %H:%M:%S',
                filename='myapp.log',
                filemode='w')'''

from openpyxl.styles import numbers, is_date_format

INCOME_PATH = r"C:\Evelyn\income"
REPORT_PATH = r"C:\Evelyn\report\2018 report"
LOG_FILE = 'claims_log.txt'
REPORT_KIND_ACU = "ACU"
REPORT_KIND_CHIRO = "CHIRO"
REPORT_KIND_PT = "PT"
REPORT_KIND_ALL= "ALL"
CHARTS_PATH=r"C:\Evelyn\charts\lexington"

def data_collection(data_path): 
    for filename in os.listdir(data_path):
        fp = os.path.join(data_path, filename)
        if os.path.isfile(fp):
            file_name_ar = fp.split("\\", fp.count("\\"))
            patient_name = file_name_ar[4]
            patient_name = patient_name[0:-5]
            patient_name = patient_name.rstrip()
            patient_name = patient_name.lstrip(",")
            patient_folder_path = os.path.join(CHARTS_PATH, patient_name)
            if os.path.exists(patient_folder_path):
                shutil.rmtree(patient_folder_path)
            #os.mkdir(patient_folder_path)
            shutil.copytree("patient_name", patient_folder_path)  
            #open(patient_intakeForm_path, "wb").write(open("patient_name", "rb").read())
           
        else:
            data_collection(fp)

def date_compare(year0,mon0,day0, year1,mon1,day1):
    year0 = int(year0)
    mon0 = int(mon0)
    day0 = int(day0)
    year1 = int(year1)
    mon1 = int(mon1)
    day1 = int(day1)
    
    
    if year0 > year1: 
        return True
    elif year0 == year1 :
        if mon0 > mon1:
            return True
        elif mon0 == mon1:
            if day0 >= day1:
                return True
            else:
                return False
        else:
            return False
    else :
        return False
    
#begin_date & end_date format [yyyymmdd]
def doing_financial_report(begin_date, end_date, report_kind = REPORT_KIND_ALL):
    acu_report_name = REPORT_PATH+"\\"+begin_date+"-"+end_date+"-"+REPORT_KIND_ACU+".xlsx"
    chiro_report_name = REPORT_PATH+"\\"+begin_date+"-"+end_date+"-"+REPORT_KIND_CHIRO+".xlsx"
    pt_report_name = REPORT_PATH+"\\"+begin_date+"-"+end_date+"-"+REPORT_KIND_PT+".xlsx"
    if os.path.exists(acu_report_name):
        os.remove(acu_report_name)
    if os.path.exists(chiro_report_name):
        os.remove(chiro_report_name)
    if os.path.exists(pt_report_name):
        os.remove(pt_report_name)
    #acu
    acu_wb = openpyxl.Workbook()
    acu_ws = acu_wb.active
    #chiro
    chiro_wb = openpyxl.Workbook()
    chiro_ws = chiro_wb.active
    #pt
    pt_wb = openpyxl.Workbook()
    pt_ws = pt_wb.active
        
    data_collection(ACCOUNTING_PATH, begin_date, end_date, acu_ws, chiro_ws, pt_ws, report_kind)

    if REPORT_KIND_ACU == report_kind:
        #acu
        acu_wb.save(acu_report_name)
    elif REPORT_KIND_CHIRO == report_kind:
        #chiro
        chiro_wb.save(chiro_report_name)
    elif REPORT_KIND_PT == report_kind:
        #pt
        pt_wb.save(pt_report_name)
    else:
        #acu
        acu_wb.save(acu_report_name)
        #chiro
        chiro_wb.save(chiro_report_name)
        #pt
        pt_wb.save(pt_report_name)
    
    

#doing_financial_report("20180224","20180226")
data_collection(INCOME_PATH)
