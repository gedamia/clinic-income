#!/usr/bin/env python
#encoding: utf-8

import importlib
import sys
import random
from urllib.request import urlopen
from urllib.request import Request
import string
import os
import re
import xlrd
import openpyxl
import sys
import os
import re
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import numbers, is_date_format
import logging
import logging.handlers
import time

LOG_FILE = 'claims_log.txt'
logging.basicConfig(level=logging.DEBUG,
                format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                datefmt='%a, %d %b %Y %H:%M:%S',
                filename='myapp.log',
                filemode='w')

AETNA_ACU_PATH = r'C:\Evelyn\function\need to do report\eob\acu\Aetna\aetna_acu_eob.xlsx'
AETNA_PT_PATH = r'C:\Evelyn\function\need to do eob\pt\eob_aetna_pt.xlsx'
CIGNA_ACU_PATH = r'C:\Evelyn\function\need to do report\eob\acu\Cigna\cigna_acu_eob.xlsx'
UHC_ACU_PATH = r'C:\Evelyn\function\need to do report\eob\acu\UHC\uhc_acu_eob.xlsx'

INCOME_PATH_TEST = r"C:\Evelyn\income"
NO_PAY_LIST = r"C:\Evelyn\function\need to do report\no_pay_list.txt"

AETNA_ACU_NAME_TAG = "Patient Name: "
AETNA_ACU_DOS_TAG = "DATES CODE "
AETNA_ACU_CHARGE_TAG = "TOTALS "
AETNA_ACU_PAID_TAG = "ISSUED AMT: "
AETNA_ACU_NO_PAY_FLAG = "NO PAY"
AETNA_ACU_TAG = "AETNA_ACU_PTN"

AETNA_PT_NAME_TAG = "Member Name:"
AETNA_PT_DOS_TAG = "DOS PL "
AETNA_PT_CHARGE_TAG = "Totals: "
AETNA_PT_PAID_TAG = "Paid Amount "
AETNA_PT_NO_PAY_FLAG = "$0.00"
AETNA_PT_TAG = "AETNA_PT_PTN"

CIGNA_ACU_NAME_TAG = "PATIENT NAME: "
CIGNA_ACU_DOS_TAG = "TOTAL "
CIGNA_ACU_CHARGE_TAG = "TOTAL "
CIGNA_ACU_PAID_TAG = "PAYMENT OF"
CIGNA_ACU_NO_PAY_FLAG = "$0.00"
CIGNA_ACU_TAG = "CIGNA_ACU_PTN"

UHC_ACU_NAME_TAG = "PATIENT: "
UHC_ACU_DOS_TAG = "PATIENT: "
UHC_ACU_CHARGE_TAG = "SUBTOTAL"
UHC_ACU_PAID_TAG = "SUBTOTAL"
UHC_ACU_TAG = "UHC_ACU_PTN"



def searching( path, last_name, first_name, charge, dos_d, dos_m, dos_y, paid, ded):
    for filename in os.listdir(path):
        fp = os.path.join(path, filename)
        if not os.path.isfile(fp):
            searching( fp, last_name, first_name, charge, dos_d, dos_m, dos_y, paid, ded)
        else:
            if bool( re.search(last_name, filename, re.IGNORECASE) and re.search(first_name, filename, re.IGNORECASE) ):
                filling( fp, charge, dos_d, dos_m, dos_y, paid, ded)
                return
            else:
                continue
    no_pay_list = open(NO_PAY_LIST, 'a')
    no_pay_list.write(last_name+","+first_name+" -----------------------------------\n\n")
    no_pay_list.close()
    return

def filling( path, charge, dos_d, dos_m, dos_y, paid, ded):
    
    data = openpyxl.load_workbook(path)
    sheet_names = data.get_sheet_names()
    sheet0 = data.get_sheet_by_name(sheet_names[0])
    for i in range(1,sheet0.max_row +1):
        data_of_service = str(sheet0.cell(row=i, column=3).value)
        charge_amt = sheet0.cell(row=i, column=7).value
        dos_date = data_of_service.split(" ")[0].split("-")
        
        if len(dos_date) > 2:
            dos_year = (data_of_service.split(" "))[0].split("-")[0]
            dos_month = (data_of_service.split(" "))[0].split("-")[1]
            dos_day = (data_of_service.split(" "))[0].split("-")[2]
            
            if(   (dos_day == dos_d) and (dos_year[-2:] == dos_y) and (dos_month == dos_m) and int(charge) == int(charge_amt) ):
                sheet0.cell(row=i, column=10).value = time.strftime("%m/%d/%Y")
                sheet0.cell(row=i, column=11).value = paid
                sheet0.cell(row=i, column=12).value = "payspan"
                if ded > 0:
                    sheet0.cell(row=i, column=14).value = "ded"+str(ded)
                
                if (AETNA_ACU_NO_PAY_FLAG == paid) or (AETNA_PT_NO_PAY_FLAG == paid):
                    
                    #no_pay_list.write("NO PAY ------------------------ "+path+","+dos_m+"/"+dos_d+"/"+dos_y+","+str(charge)+"\n\n")
                    print("no pay"+path+","+dos_m+"/"+dos_d+"/"+dos_y+","+str(charge) +","+str(ded)+"\n\n")
                break
    data.save(path)
            
def parsing_aetna_acu(income_path, eob_path):
    first_name = ""
    last_name = ""
    charge = 0
    dos_d = ""
    dos_m = ""
    dos_y = ""
    paid = 0
    eob_data = xlrd.open_workbook(eob_path)
    eob_table = eob_data.sheets()[0]
    for i in range(eob_table.nrows):
        data_value = str(eob_table.cell(i,0).value)
        if data_value.startswith(AETNA_ACU_NAME_TAG):       #name
            patient_name = data_value.split(":", data_value.count(":"))
            patient_name = patient_name[1].split("(", patient_name[1].count("("))
            patient_name = patient_name[0].split(" ", patient_name[0].count(" "))
            first_name = patient_name[1]
            last_name = patient_name[-2]
        elif data_value.startswith(AETNA_ACU_DOS_TAG):          #dos
            data_value = str(eob_table.cell(i+1,0).value)
            dos = data_value.split(" ", data_value.count(" "))
            if len(dos) > 2:
                dos = str(dos[0]).split("/", str(dos[0]).count("/"))
                dos_d = dos[1]
                dos_m = dos[0]
                dos_y = dos[2]
        elif data_value.startswith(AETNA_ACU_CHARGE_TAG):       #charge
            charge = data_value.split(" ", data_value.count(" "))
            charge = charge[1].replace(",","") #cool method!
            charge = float(charge)
        elif data_value.startswith(AETNA_ACU_PAID_TAG):         #paid
            paid = data_value.split(":", data_value.count(":"))
            #begin to write into patients' claims
            searching( income_path, last_name, first_name, charge, dos_d, dos_m, dos_y, paid[1][1:])
            #end to write into patients' claims
        else:
            continue

def parsing_aetna_pt(income_path, eob_path):
    first_name = ""
    last_name = ""
    charge = 0
    dos_d = ""
    dos_m = ""
    dos_y = ""
    paid = 0
    eob_data = xlrd.open_workbook(eob_path)
    eob_table = eob_data.sheets()[0]
    for i in range(eob_table.nrows):
        data_value = str(eob_table.cell(i,0).value)
        if data_value.startswith(AETNA_PT_NAME_TAG):       #name
            patient_name = data_value.split(":", data_value.count(":"))
            patient_name = patient_name[1].split(" Product Type: ", patient_name[1].count(" Product Type: "))
            patient_name = patient_name[0].split(" ", patient_name[0].count(" "))
            first_name = patient_name[1]
            last_name = patient_name[-3]
        elif data_value.startswith(AETNA_PT_DOS_TAG):          #dos
            data_value = str(eob_table.cell(i+1,0).value)
            dos = data_value.split(" ", data_value.count(" "))
            if len(dos) > 2:
                dos = str(dos[0]).split("/", str(dos[0]).count("/"))
                dos_d = dos[1]
                dos_m = dos[0]
                dos_y = dos[2][-2:]
        elif data_value.startswith(AETNA_PT_CHARGE_TAG):       #charge
            charge = data_value.split(" ", data_value.count(" "))
            deductable = charge[4].replace(",","") #cool method!
            deductable = deductable.replace("(","")
            deductable = deductable.replace(")","")
            deductable = float(deductable[1:])


            
            charge = charge[1].replace(",","") #cool method!
            charge = charge.replace("(","")
            charge = charge.replace(")","")
            charge = float(charge[1:])

        elif data_value.startswith(AETNA_PT_PAID_TAG):         #paid
            paid = data_value.split(" ", data_value.count(" "))
            #begin to write into patients' claims
            searching( income_path, last_name, first_name, charge, dos_d, dos_m, dos_y, paid[2], deductable)
            #end to write into patients' claims
        else:
            continue
    return
def parsing_cigna_acu(income_path, eob_path):
    first_name = ""
    last_name = ""
    charge = 0
    dos_d = ""
    dos_m = ""
    dos_y = ""
    paid = 0
    eob_data = xlrd.open_workbook(eob_path)
    eob_table = eob_data.sheets()[0]
    for i in range(eob_table.nrows):
        data_value = str(eob_table.cell(i,0).value)
        if data_value.startswith(CIGNA_ACU_NAME_TAG):       #name
            patient_name = data_value.split(":", data_value.count(":"))
            patient_name = patient_name[1].split("PATIENT#:", patient_name[1].count("PATIENT#:"))
            patient_name = patient_name[0].split(" ", patient_name[0].count(" "))
            first_name = patient_name[1]
            last_name = patient_name[-2]
            
        elif (data_value.startswith(CIGNA_ACU_DOS_TAG)) or (data_value.startswith(CIGNA_ACU_CHARGE_TAG)):          #dos or charge
            data_value = str(eob_table.cell(i-1,0).value)
            dos = data_value[3:19]
            dos = dos.replace(" ","")
            dos_m = dos[:2]
            dos_d = dos[2:4]
            dos_y = dos[6:]

            data_value = str(eob_table.cell(i,0).value)
            charge = data_value.split(".", data_value.count("."))
            charge = charge[0]+"."+charge[1][0:4]
            charge = charge[6:].replace(",","") #cool method!
            charge = charge.replace(" ","")
            charge = float(charge)
            
        elif data_value.startswith(CIGNA_ACU_PAID_TAG):         #paid
            paid = data_value.split(" ", data_value.count(" "))[2]
            #begin to write into patients' claims
            searching( income_path, last_name, first_name, charge, dos_d, dos_m, dos_y, paid)
            #end to write into patients' claims
        else:
            continue
    return
def parsing_uhc_acu(income_path, eob_path):
    first_name = ""
    last_name = ""
    charge = 0
    dos_d = ""
    dos_m = ""
    dos_y = ""
    paid = 0
    eob_data = xlrd.open_workbook(eob_path)
    eob_table = eob_data.sheets()[0]
    for i in range(eob_table.nrows):
        data_value = str(eob_table.cell(i,0).value)
        
        if (data_value.startswith(UHC_ACU_NAME_TAG)) or (data_value.startswith(UHC_ACU_DOS_TAG)):       #name or dos

            patient_name = data_value.split(":", data_value.count(":"))
            patient_name = patient_name[1].split("(", patient_name[1].count("("))
            patient_name = patient_name[0].split(" ", patient_name[0].count(" "))
            first_name = patient_name[1]
            last_name = patient_name[-2]
            

            dos_value = str(eob_table.cell(i+5,0).value)
            dos = dos_value.replace(" ","")
            dos = dos.split("-")[0]
            dos_m = dos[:2]
            dos_d = dos[3:5]
            dos_y = dos[6:]
            
        elif (re.search(UHC_ACU_CHARGE_TAG, data_value)) or (data_value.startswith(UHC_ACU_PAID_TAG)):  #charge or paid
            charge = data_value.split(UHC_ACU_CHARGE_TAG, data_value.count(UHC_ACU_CHARGE_TAG))
            charge = charge[1].split(" ", charge[1].count(" "))
            charge = charge[1][1:]
            charge = float(charge.replace(",",""))
            
      
            paid = data_value.split(UHC_ACU_CHARGE_TAG, data_value.count(UHC_ACU_CHARGE_TAG))
            paid = paid[1].split(" ", paid[1].count(" "))
            #paid = paid[3][1:]
            #paid = float(paid.replace(",",""))
            if (paid[-1].startswith("$")):
                paid = paid[-1]
            else:
                paid = paid[-2]
            
            #begin to write into patients' claims
            searching( income_path, last_name, first_name, charge, dos_d, dos_m, dos_y, paid)
            #end to write into patients' claims
        else:
            continue
    return



#parsing_aetna_acu(INCOME_PATH_TEST, AETNA_ACU_PATH)
parsing_aetna_pt(INCOME_PATH_TEST, AETNA_PT_PATH)
#parsing_cigna_acu(INCOME_PATH_TEST, CIGNA_ACU_PATH)
#parsing_uhc_acu(INCOME_PATH_TEST, UHC_ACU_PATH)
  




