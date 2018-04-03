import xlrd
import openpyxl
import sys
import os
import re
import time
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
import uuid

import logging
import logging.handlers 
'''logging.basicConfig(level=logging.DEBUG,
                format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                datefmt='%a, %d %b %Y %H:%M:%S',
                filename='myapp.log',
                filemode='w')'''

from openpyxl.styles import numbers, is_date_format

ACCOUNTING_PATH = r"C:\Evelyn\accounting"
REPORT_PATH = r"C:\Evelyn\report\2018 report"
LOG_FILE = 'claims_log.txt'
REPORT_KIND_ACU = "ACU"
REPORT_KIND_CHIRO = "CHIRO"
REPORT_KIND_PT = "PT"
REPORT_KIND_ALL= "ALL"

def data_collection(data_path, begin_date, end_date, report_ws_acu, report_ws_chiro, report_ws_pt,report_kind = REPORT_KIND_ALL,): 
    for filename in os.listdir(data_path):
        fp = os.path.join(data_path, filename)
        if os.path.isfile(fp):
            data = xlrd.open_workbook(fp)
            data_table = data.sheets()[0]

            for i in range(data_table.nrows):
                if data_table.ncols < 10:
                    continue
                record_date = data_table.cell(i, 9).value
                record_bill_amount = data_table.cell(i, 6).value
                if record_date != "" and  isinstance(record_date, float):
                    [r_year,r_month,r_day,r_hour,r_minitue,r_second] = xlrd.xldate_as_tuple(record_date, 0)
                    if (date_compare(r_year, r_month, r_day, begin_date[0:4], begin_date[4:6], begin_date[6:]) and
                        date_compare(end_date[0:4],end_date[4:6],end_date[6:], r_year, r_month, r_day) ):
                        report_ws_acu.append(data_table.row_values(i))
                        report_ws_chiro.append(data_table.row_values(i))
                        report_ws_pt.append(data_table.row_values(i))
                    else:
                        continue
                else:
                    continue
        else:
            data_collection(fp, begin_date, end_date, report_ws_acu, report_ws_chiro, report_ws_pt, report_kind)

def date_compare(year0,mon0,day0, year1,mon1,day1):
    year0 = int(year0)
    mon0 = int(mon0)
    day0 = int(day0)
    year1 = int(year1)
    mon1 = int(mon1)
    day1 = int(day1)
    
    
    if year0 > year1: 
        return True
    elif year0 == year1 :
        if mon0 > mon1:
            return True
        elif mon0 == mon1:
            if day0 >= day1:
                return True
            else:
                return False
        else:
            return False
    else :
        return False
    
#begin_date & end_date format [yyyymmdd]
def doing_financial_report(begin_date, end_date, report_kind = REPORT_KIND_ALL):
    acu_report_name = REPORT_PATH+"\\"+begin_date+"-"+end_date+"-"+REPORT_KIND_ACU+".xlsx"
    chiro_report_name = REPORT_PATH+"\\"+begin_date+"-"+end_date+"-"+REPORT_KIND_CHIRO+".xlsx"
    pt_report_name = REPORT_PATH+"\\"+begin_date+"-"+end_date+"-"+REPORT_KIND_PT+".xlsx"
    if os.path.exists(acu_report_name):
        os.remove(acu_report_name)
    if os.path.exists(chiro_report_name):
        os.remove(chiro_report_name)
    if os.path.exists(pt_report_name):
        os.remove(pt_report_name)
    #acu
    acu_wb = openpyxl.Workbook()
    acu_ws = acu_wb.active
    #chiro
    chiro_wb = openpyxl.Workbook()
    chiro_ws = chiro_wb.active
    #pt
    pt_wb = openpyxl.Workbook()
    pt_ws = pt_wb.active
        
    data_collection(ACCOUNTING_PATH, begin_date, end_date, acu_ws, chiro_ws, pt_ws, report_kind)

    if REPORT_KIND_ACU == report_kind:
        #acu
        acu_wb.save(acu_report_name)
    elif REPORT_KIND_CHIRO == report_kind:
        #chiro
        chiro_wb.save(chiro_report_name)
    elif REPORT_KIND_PT == report_kind:
        #pt
        pt_wb.save(pt_report_name)
    else:
        #acu
        acu_wb.save(acu_report_name)
        #chiro
        chiro_wb.save(chiro_report_name)
        #pt
        pt_wb.save(pt_report_name)
    
    
if '005056c00008' == uuid.UUID(int = uuid.getnode()).hex[-12:] :
    doing_financial_report("20180224","20180226")
